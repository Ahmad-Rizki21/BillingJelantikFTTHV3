import pandas as pd
from typing import List
import openpyxl
from io import BytesIO
from typing import Optional
from datetime import datetime, timedelta, date, timezone
import io
import csv
import chardet  # Add this import for encoding detection
import json
import uuid
from typing import Union, TYPE_CHECKING

from fastapi import APIRouter, Depends, HTTPException, status, Request, Header
from fastapi.responses import StreamingResponse, Response
from dateutil import parser
from dateutil.relativedelta import relativedelta

from sqlalchemy import func, or_, and_
from ..models.user import User as UserModel
from ..models.role import Role as RoleModel
from ..websocket_manager import manager

# Import has_permission function
from ..auth import has_permission, get_current_active_user

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import joinedload, selectinload
from pydantic import ValidationError
import logging

from ..models.invoice import Invoice as InvoiceModel
from ..models.langganan import Langganan as LanggananModel
from ..models.pelanggan import Pelanggan as PelangganModel
from ..schemas.invoice import (
    Invoice as InvoiceSchema,
    InvoiceGenerate,
    MarkAsPaidRequest,
)
from ..database import get_db

from sqlalchemy import select, func
from ..services import mikrotik_service
from ..config import settings
from ..services import xendit_service, mikrotik_service
from ..services.payment_callback_service import check_duplicate_callback, log_callback_processing
from ..services.rate_limiter import create_invoice_with_rate_limit, InvoicePriority

# Import our logging utilities
from ..logging_utils import sanitize_log_data

logger = logging.getLogger("app.routers.invoice")


# Helper functions untuk safe date conversion
def safe_to_datetime(date_obj) -> datetime:
    """Convert date/datetime ke datetime dengan aman."""
    if date_obj is None:
        return datetime.now(timezone.utc)

    if isinstance(date_obj, datetime):
        return date_obj

    # Handle SQLAlchemy Date object atau Python date
    try:
        if hasattr(date_obj, "strftime"):
            return datetime.combine(date_obj, datetime.min.time())
        else:
            # Fallback untuk SQLAlchemy Date
            return datetime.combine(date_obj, datetime.min.time())
    except (AttributeError, TypeError):
        return datetime.now(timezone.utc)


def safe_format_date(date_obj, format_str: str = "%Y-%m-%d") -> str:
    """Format date dengan aman."""
    if date_obj is None:
        return ""

    try:
        if hasattr(date_obj, "strftime"):
            return date_obj.strftime(format_str)
        else:
            # Handle SQLAlchemy Date
            return str(date_obj)
    except (AttributeError, TypeError):
        return str(date_obj) if date_obj else ""


def safe_get_day(date_obj) -> int:
    """Get day dari date dengan aman."""
    if date_obj is None:
        return 1

    try:
        if hasattr(date_obj, "day"):
            return date_obj.day
        else:
            # Handle SQLAlchemy Date - convert dulu
            dt = safe_to_datetime(date_obj)
            return dt.day
    except (AttributeError, TypeError):
        return 1


def safe_relativedelta_operation(date_obj, delta_months: int):
    """Safe operation untuk relativedelta dengan date/datetime."""
    dt = safe_to_datetime(date_obj)
    return dt + relativedelta(months=delta_months)


router = APIRouter(
    prefix="/invoices",
    tags=["Invoices"],
    responses={404: {"description": "Not found"}},
)


# GET /invoices - Ambil semua data invoice
# Buat nampilin list invoice dengan fitur filter dan pencarian lengkap
# Query parameters:
# - search: cari berdasarkan nomor invoice, nama pelanggan, atau ID pelanggan
# - status_invoice: filter berdasarkan status (Belum Dibayar, Lunas, Kadaluarsa)
# - start_date: filter tanggal jatuh tempo mulai dari
# - end_date: filter tanggal jatuh tempo sampai dengan
# - show_active_only: kalo true, cuma tampilkan invoice yang punya payment_link aktif (90 hari terakhir)
# - skip: offset pagination (default: 0)
# - limit: jumlah data per halaman (default: semua)
# Response: list invoice dengan relasi lengkap (pelanggan, langganan, paket, data teknis)
# Performance: eager loading biar ga N+1 query
@router.get("/", response_model=List[InvoiceSchema])
async def get_all_invoices(
    db: AsyncSession = Depends(get_db),
    search: Optional[str] = None,
    status_invoice: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    show_active_only: Optional[bool] = False,  # <-- FILTER untuk link pembayaran aktif saja
    skip: int = 0,  # <-- TAMBAHKAN INI
    limit: Optional[int] = None,  # <-- TAMBAHKAN INI
):
    """Mengambil semua data invoice dengan filter."""
    # OPTIMIZED: Query dengan comprehensive eager loading untuk mencegah semua N+1 problems
    query = (
        select(InvoiceModel).join(InvoiceModel.pelanggan)
        # PERBAIKAN: Eager load semua relasi yang sering diakses bersama invoice
        # untuk mencegah N+1 queries saat data di-serialize ke response
        .options(
            joinedload(InvoiceModel.pelanggan).options(
                joinedload(PelangganModel.harga_layanan),
                joinedload(PelangganModel.data_teknis),
                joinedload(PelangganModel.langganan).joinedload(LanggananModel.paket_layanan),
            )
        )
    )

    if search:
        search_term = f"%{search}%"
        query = query.where(
            or_(
                InvoiceModel.invoice_number.ilike(search_term),
                PelangganModel.nama.ilike(search_term),
                InvoiceModel.id_pelanggan.ilike(search_term),
            )
        )

    if status_invoice:
        query = query.where(InvoiceModel.status_invoice == status_invoice)

    if start_date:
        query = query.where(InvoiceModel.tgl_jatuh_tempo >= start_date)
    if end_date:
        query = query.where(InvoiceModel.tgl_jatuh_tempo <= end_date)

    # Filter untuk hanya menampilkan invoice dengan link pembayaran aktif
    if show_active_only:
        # Hanya tampilkan invoice yang:
        # 1. Memiliki payment_link (baik yang sudah lunas maupun belum)
        # 2. Link pembayaran masih relevan (tidak terlalu tua)
        from datetime import date, timedelta

        cutoff_date = date.today() - timedelta(days=90)  # Invoice 90 hari terakhir

        query = query.where(
            and_(
                InvoiceModel.payment_link.isnot(None), InvoiceModel.payment_link != "", InvoiceModel.tgl_invoice >= cutoff_date
            )
        ).order_by(InvoiceModel.tgl_invoice.desc())

    # Terapkan paginasi setelah semua filter
    if limit is not None:
        query = query.offset(skip).limit(limit)
    # ---------------------------

    result = await db.execute(query)
    # FIX: Tambahkan .unique() untuk collection eager loading
    return result.scalars().unique().all()


def parse_xendit_datetime(iso_datetime_str: str) -> datetime:
    """Fungsi untuk mengkonversi format datetime ISO 8601 dari Xendit."""
    try:
        if not iso_datetime_str:
            pass  # This line was misplaced and caused an IndentationError. It's removed as it didn't have a clear purpose here.
        if iso_datetime_str.endswith("Z"):
            iso_datetime_str = iso_datetime_str[:-1] + "+00:00"
        return datetime.fromisoformat(iso_datetime_str)
    except (ValueError, TypeError):
        return datetime.now(timezone.utc)


async def _process_successful_payment(db: AsyncSession, invoice: InvoiceModel, payload: dict | None = None):
    """Fungsi terpusat untuk menangani logika setelah invoice lunas."""

    pelanggan = invoice.pelanggan
    if not pelanggan or not pelanggan.langganan:
        logger.error(f"Pelanggan atau langganan tidak ditemukan untuk invoice {invoice.invoice_number}")
        return

    langganan = pelanggan.langganan[0]

    # Cek apakah langganan sebelumnya berstatus 'Suspended'
    is_suspended_or_inactive = langganan.status == "Suspended" or not langganan.status

    # Update status invoice (tetap)
    invoice.status_invoice = "Lunas"
    if payload:
        invoice.paid_amount = float(payload.get("paid_amount", invoice.total_harga or 0))
        paid_at_str = payload.get("paid_at")
        invoice.paid_at = parse_xendit_datetime(paid_at_str) if paid_at_str else datetime.now(timezone.utc)
    else:
        invoice.paid_amount = invoice.total_harga
        invoice.paid_at = datetime.now(timezone.utc)

    db.add(invoice)

    next_due_date = None

    if langganan.metode_pembayaran == "Prorate":
        paket = langganan.paket_layanan
        brand = pelanggan.harga_layanan
        langganan.metode_pembayaran = "Otomatis"
        current_due_date = invoice.tgl_jatuh_tempo

        if not paket or not brand:
            logger.error(f"Data paket/brand tidak lengkap untuk langganan ID {langganan.id}")
            # Set fallback jika data tidak ada, agar tidak crash
            current_due_datetime = safe_to_datetime(current_due_date)
            next_due_datetime = current_due_datetime + relativedelta(months=1)
            next_due_date = next_due_datetime.date().replace(day=1)
        else:
            # Hitung harga normal penuh sebagai pembanding
            harga_paket = float(paket.harga)
            pajak_persen = float(brand.pajak)
            harga_normal_full = harga_paket * (1 + (pajak_persen / 100))

            # Logika Pembeda: Apakah ini tagihan prorate biasa atau gabungan?
            if float(invoice.total_harga or 0) > (harga_normal_full + 1):
                # Skenario 1: INI ADALAH TAGIHAN GABUNGAN
                # Convert Date to datetime untuk relativedelta - handle SQLAlchemy Date
                current_due_datetime = safe_to_datetime(current_due_date)
                next_due_datetime = current_due_datetime + relativedelta(months=2)
                next_due_date = next_due_datetime.date().replace(day=1)
                logger.info(f"Tagihan gabungan terdeteksi. Jatuh tempo berikutnya diatur ke {next_due_date}")
            else:
                # Skenario 2: INI ADALAH TAGIHAN PRORATE BIASA
                # Convert Date to datetime untuk relativedelta - handle SQLAlchemy Date
                current_due_datetime = safe_to_datetime(current_due_date)
                next_due_datetime = current_due_datetime + relativedelta(months=1)
                next_due_date = next_due_datetime.date().replace(day=1)
                logger.info(f"Tagihan prorate biasa terdeteksi. Jatuh tempo berikutnya diatur ke {next_due_date}")

            # Reset harga langganan ke harga normal untuk bulan-bulan berikutnya
            langganan.harga_awal = round(harga_normal_full, 0)

    else:  # Skenario 3: Jika sudah Otomatis (PEMBAYARAN BULANAN NORMAL)
        current_due_date = langganan.tgl_jatuh_tempo or date.today()
        # Convert Date to datetime untuk relativedelta - handle SQLAlchemy Date
        current_due_datetime = safe_to_datetime(current_due_date)
        next_due_datetime = current_due_datetime + relativedelta(months=1)
        next_due_date = next_due_datetime.date().replace(day=1)

    # Update langganan (tidak berubah)
    langganan.status = "Aktif"
    langganan.tgl_jatuh_tempo = next_due_date
    langganan.tgl_invoice_terakhir = date.today()
    db.add(langganan)

    # HANYA trigger update Mikrotik jika status sebelumnya adalah 'Suspended'
    if is_suspended_or_inactive:
        data_teknis = pelanggan.data_teknis
        if not data_teknis:
            logger.error(f"Data Teknis tidak ditemukan untuk pelanggan ID {pelanggan.id}. Mikrotik update dilewati.")
        else:
            try:
                # Panggil fungsi dengan SEMUA argumen yang dibutuhkan
                await mikrotik_service.trigger_mikrotik_update(
                    db,
                    langganan,
                    data_teknis,
                    data_teknis.id_pelanggan,  # old_id_pelanggan diisi dengan id saat ini
                )
                logger.info(f"Berhasil trigger re-aktivasi Mikrotik untuk langganan ID {langganan.id}")

                # Jika berhasil, set flag pending sync (jika ada) kembali ke False
                if data_teknis.mikrotik_sync_pending:
                    data_teknis.mikrotik_sync_pending = False
                    db.add(data_teknis)

            except Exception as e:
                # Jika GAGAL, catat error DAN set flag retry menjadi True
                logger.error(
                    f"Gagal trigger re-aktivasi Mikrotik untuk langganan ID {langganan.id}: {e}. Menandai untuk dicoba lagi."
                )
                data_teknis.mikrotik_sync_pending = True
                db.add(data_teknis)
    else:
        logger.info(f"Langganan ID {langganan.id} sudah Aktif. Mikrotik update dilewati.")

    # Notif ke frontend
    try:
        target_roles = ["Admin", "NOC", "Finance"]
        query = select(UserModel.id).join(RoleModel).where(func.lower(RoleModel.name).in_([r.lower() for r in target_roles]))
        result = await db.execute(query)
        target_user_ids = result.scalars().all()

        if target_user_ids:
            # Pastikan pelanggan sudah di-load dengan benar
            pelanggan_nama = pelanggan.nama if pelanggan else "N/A"
            notification_payload = {
                "type": "new_payment",
                "message": f"Pembayaran untuk invoice {invoice.invoice_number} dari {pelanggan_nama} telah diterima.",
                "timestamp": datetime.now().isoformat(),
                "data": {
                    "invoice_id": invoice.id,
                    "invoice_number": invoice.invoice_number,
                    "pelanggan_nama": pelanggan_nama,
                    "amount": (float(invoice.total_harga) if invoice.total_harga else 0.0),
                    "payment_method": invoice.metode_pembayaran or "Unknown",
                    "timestamp": datetime.now().isoformat(),
                },
            }
            # Tambahkan log ini untuk memastikan user ID ditemukan
            logger.info(f"Mencoba mengirim notifikasi pembayaran ke user IDs: {target_user_ids}")
            # Convert ke list untuk broadcast_to_roles
            user_ids_list = list(target_user_ids)
            await manager.broadcast_to_roles(notification_payload, user_ids_list)
            logger.info(f"Notifikasi pembayaran berhasil dikirim untuk invoice {invoice.invoice_number}")
        else:
            logger.warning(f"Tidak ada user dengan role Admin/CS yang ditemukan untuk dikirimi notifikasi.")

    except Exception as e:
        # 🛡️ Graceful degradation: Payment processed but notification failed
        logger.error(f"⚠️ Payment successful but notification failed for invoice {invoice.invoice_number}: {e}", exc_info=True)
        # Continue processing - payment is still valid even if notification fails

    logger.info(f"Payment processed successfully for invoice {invoice.invoice_number}")


# @router.post("/xendit-callback", status_code=status.HTTP_200_OK)
# async def handle_xendit_callback(
#     request: Request,
#     x_callback_token: Optional[str] = Header(None),
#     db: AsyncSession = Depends(get_db),
# ):
#     payload = await request.json()
#     logger.info(f"Xendit callback received. Payload: {json.dumps(payload, indent=2)}")

#     external_id = payload.get("external_id")
#     if not external_id:
#         raise HTTPException(
#             status_code=400, detail="External ID tidak ditemukan di payload"
#         )

#     try:
#         brand_prefix = external_id.split("/")[0]
#     except IndexError:
#         raise HTTPException(status_code=400, detail="Format external_id tidak valid")

#     correct_token = None
#     if brand_prefix.lower() in ["jakinet", "nagrak"]:
#         correct_token = settings.XENDIT_CALLBACK_TOKENS.get("ARTACOM")
#         logger.info("Validating with ARTACOM callback token.")
#     elif brand_prefix.lower() == "jelantik":
#         correct_token = settings.XENDIT_CALLBACK_TOKENS.get("JELANTIK")
#         logger.info("Validating with JELANTIK callback token.")

#     if not correct_token or x_callback_token != correct_token:
#         logger.warning(f"Invalid callback token received for brand '{brand_prefix}'.")
#         raise HTTPException(
#             status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid callback token"
#         )


# POST /invoices/xendit-callback - Callback dari Xendit payment gateway
# Endpoint buat terima callback dari Xendit setelah customer bayar invoice
# Request headers:
# - x_callback_token: token untuk validasi callback (sesuai brand)
# Request body: payment data dari Xendit (JSON)
# Response: success message
# Security:
# - Validasi callback token berdasarkan brand prefix
# - Cek duplikasi callback biar idempotent
# - Log semua callback untuk audit trail
# Fitur:
# - Update status invoice jadi "Lunas" kalo payment sukses
# - Update status jadi "Kadaluarsa" kalo payment expired
# - Trigger re-aktivasi Mikrotik kalo pelanggan sebelumnya suspended
# - Kirim notifikasi ke Admin/NOC/Finance
# Error handling: graceful degradation, payment tetep valid walau notifikasi gagal
@router.post("/xendit-callback", status_code=status.HTTP_200_OK)
async def handle_xendit_callback(
    request: Request,
    x_callback_token: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_db),
):
    # Get raw body for logging but filter sensitive data
    raw_body = await request.body()

    # Log the callback with filtered data
    try:
        if raw_body:
            body_str = raw_body.decode("utf-8")
            filtered_body = sanitize_log_data(body_str)
            logger.info(f"Xendit callback received. Filtered Payload: {filtered_body}")
        else:
            logger.info("Xendit callback received. Payload: Empty body")
    except Exception as e:
        logger.info(f"Xendit callback received. Payload: ***REDACTED*** (Error processing body: {str(e)})")

    payload = await request.json()
    # Log the JSON payload with sensitive data filtered
    filtered_payload = sanitize_log_data(payload)
    logger.info(f"Xendit callback received. Filtered JSON Payload: {json.dumps(filtered_payload, indent=2)}")

    # Extract IDs from payload
    xendit_id = payload.get("id")  # Xendit internal ID
    external_id = payload.get("external_id")
    xendit_status = payload.get("status")

    # Extract idempotency key if provided in headers
    idempotency_key = request.headers.get("x-idempotency-key", request.headers.get("idempotency-key"))

    if not external_id:
        raise HTTPException(status_code=400, detail="External ID tidak ditemukan di payload")

    # Check for duplicate callback
    is_duplicate = await check_duplicate_callback(db, xendit_id or external_id, external_id, idempotency_key or "")
    if is_duplicate:
        logger.info(f"Duplicated callback received and ignored: xendit_id={xendit_id}, external_id={external_id}")
        return {"message": "Callback already processed"}

    brand_prefix = None
    try:
        if "/" in external_id:
            brand_prefix = external_id.split("/")[0]
        # Untuk webhook test, tidak ada prefix brand. Kita validasi token saja
    except IndexError:
        raise HTTPException(status_code=400, detail="Format external_id tidak valid")

    # VALIDASI TOKEN SECARA LANGSUNG
    # 1. Coba validasi dengan token ARTACOMINDO (Jakinet, Nagrak)
    artacom_token = settings.XENDIT_CALLBACK_TOKENS.get("ARTACOMINDO")
    if artacom_token and x_callback_token == artacom_token:
        logger.info("Validating with ARTACOMINDO callback token.")
        # Cek apakah brand_prefix (jika ada) sesuai dengan token ini
        if brand_prefix and brand_prefix.lower() not in [
            "jakinet",
            "nagrak",
            "artacom",
        ]:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid brand for this token",
            )
        correct_token = artacom_token

    # 2. Coba validasi dengan token JELANTIK
    jelantik_token = settings.XENDIT_CALLBACK_TOKENS.get("JELANTIK")
    if jelantik_token and x_callback_token == jelantik_token:
        logger.info("Validating with JELANTIK callback token.")
        # Cek apakah brand_prefix (jika ada) sesuai dengan token ini
        if brand_prefix and brand_prefix.lower() != "jelantik":
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid brand for this token",
            )
        correct_token = jelantik_token

    # Jika tidak ada token yang cocok, kembalikan 401
    if not correct_token:
        logger.warning("Invalid callback token received.")
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid callback token")

    # Log callback processing to prevent duplicates
    logged = await log_callback_processing(
        db,
        xendit_id or external_id,  # Use xendit_id if available, otherwise use external_id
        external_id,
        xendit_status,
        payload,
        idempotency_key or "",
    )

    # If logging failed, it means another process already handled this callback
    if not logged:
        logger.info(
            f"Callback already processed by another concurrent request: xendit_id={xendit_id}, external_id={external_id}"
        )
        return {"message": "Callback already processed"}

    # Proceed with normal processing
    # Buat filter conditions untuk query invoice
    # Cari berdasarkan xendit_external_id terlebih dahulu, kemudian fallback ke invoice_number
    filter_conditions = [InvoiceModel.xendit_external_id == external_id, InvoiceModel.invoice_number == external_id]

    # Logging tambahan untuk debugging
    logger.info(f"Searching for invoice with external_id: {external_id}")

    # Optimasi query dengan menggunakan joinedload untuk relasi yang sering digunakan bersama
    # Ini akan menghindari N+1 query problem
    stmt = (
        select(InvoiceModel)
        .join(InvoiceModel.pelanggan)
        .options(
            joinedload(InvoiceModel.pelanggan).options(
                joinedload(PelangganModel.harga_layanan),
                joinedload(PelangganModel.langganan).joinedload(LanggananModel.paket_layanan),
                joinedload(PelangganModel.data_teknis),
            )
        )
        .where(or_(*filter_conditions))
    )
    invoice = (await db.execute(stmt)).unique().scalar_one_or_none()

    # Logging tambahan untuk debugging
    if not invoice:
        # Coba cari dengan LIKE untuk melihat apakah ada invoice yang mirip
        search_stmt = select(InvoiceModel).where(
            or_(InvoiceModel.invoice_number.like(f"%{external_id}%"), InvoiceModel.xendit_external_id.like(f"%{external_id}%"))
        )
        similar_invoices = (await db.execute(search_stmt)).scalars().all()
        if similar_invoices:
            logger.info(f"Found similar invoices: {[inv.invoice_number for inv in similar_invoices]}")
        else:
            logger.info("No similar invoices found")

    if not invoice:
        logger.warning(f"Invoice with external_id {external_id} not found, but callback is valid.")
        return {"message": "Callback valid, invoice not found."}

    if invoice.status_invoice == "Lunas":
        # Still check if this callback was already logged for idempotency tracking
        # This can happen if the invoice status was updated but the callback logging failed
        callback_exists = await check_duplicate_callback(db, xendit_id or external_id, external_id, idempotency_key or "")
        if not callback_exists:
            # If no callback record exists, create one for tracking purposes
            await log_callback_processing(
                db,
                xendit_id or external_id,  # Use xendit_id if available, otherwise use external_id
                external_id,
                xendit_status,
                payload,
                idempotency_key or "",
            )

        logger.info(f"Invoice {invoice.invoice_number} already has status Lunas, callback ignored.")
        return {"message": "Invoice already processed"}

    try:
        if xendit_status == "PAID":
            await _process_successful_payment(db, invoice, payload)

        elif xendit_status == "EXPIRED":
            invoice.status_invoice = "Kadaluarsa"
            db.add(invoice)

        await db.commit()
    except Exception as e:
        await db.rollback()
        logger.error(f"Error processing Xendit callback for external_id {external_id}: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error while processing callback.")

    return {"message": "Callback processed successfully"}


# POST /invoices/generate - Generate invoice manual
# Buat bikin invoice baru secara manual berdasarkan langganan yang udah ada
# Request body: langganan_id
# Response: data invoice yang baru dibuat dengan payment link dari Xendit
# Permission: butuh permission "create_invoices"
# Fitur:
# - Hitung harga otomatis berdasarkan paket dan pajak
# - Generate nomor invoice otomatis
# - Create payment link di Xendit
# - Dynamic description (prorate/bulan penuh)
# - Format nomor telepon buat Xendit (0xx -> 62xx)
# Validation:
# - Cek langganan harus ada
# - Cek pelanggan harus ada
# - Cek status langganan (boleh: Aktif, Suspended)
# - Cek duplicate invoice di bulan yang sama
# Error handling: rollback transaction kalo ada error
@router.post(
    "/generate",
    response_model=InvoiceSchema,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(has_permission("create_invoices"))],
)
async def generate_manual_invoice(invoice_data: InvoiceGenerate, db: AsyncSession = Depends(get_db)):
    """Membuat satu invoice secara manual berdasarkan langganan_id."""
    # PERBAIKAN: Query dengan eager loading yang lebih robust untuk mencegah N+1 problems
    stmt = (
        select(LanggananModel)
        .where(LanggananModel.id == invoice_data.langganan_id)
        .options(
            joinedload(LanggananModel.pelanggan).options(
                joinedload(PelangganModel.harga_layanan),
                joinedload(PelangganModel.data_teknis),
            ),
            joinedload(LanggananModel.paket_layanan),
        )
    )
    result = await db.execute(stmt)
    langganan = result.unique().scalar_one_or_none()

    if not langganan:
        raise HTTPException(status_code=404, detail=f"Langganan dengan ID {invoice_data.langganan_id} tidak ditemukan")

    # VALIDASI: Periksa apakah pelanggan dari langganan benar-benar ada
    if not langganan.pelanggan:
        logger.error(
            f"Langganan ID {langganan.id} memiliki pelanggan_id {langganan.pelanggan_id} tapi data pelanggan tidak ditemukan di database"
        )
        raise HTTPException(
            status_code=404,
            detail=f"Pelanggan dengan ID {langganan.pelanggan_id} tidak ditemukan untuk langganan ID {langganan.id}. Data mungkin tidak konsisten.",
        )

    if langganan.status == "Berhenti":
        raise HTTPException(
            status_code=400,
            detail=f"Gagal membuat invoice. Status langganan untuk pelanggan '{langganan.pelanggan.nama}' adalah 'Berhenti'.",
        )

    pelanggan = langganan.pelanggan
    paket = langganan.paket_layanan
    if not pelanggan or not paket or not pelanggan.harga_layanan or not pelanggan.data_teknis:
        missing_data = []
        if not pelanggan:
            missing_data.append("pelanggan")
        if not paket:
            missing_data.append("paket_layanan")
        if not pelanggan.harga_layanan:
            missing_data.append("harga_layanan/brand")
        if not pelanggan.data_teknis:
            missing_data.append("data_teknis")

        raise HTTPException(
            status_code=400,
            detail=f"Data pendukung tidak lengkap untuk langganan ID {langganan.id}: {', '.join(missing_data)}",
        )

    brand = pelanggan.harga_layanan
    data_teknis = pelanggan.data_teknis

    if not paket.harga or not brand.pajak:
        raise HTTPException(status_code=400, detail="Harga paket atau pajak tidak valid")

    existing_invoice_stmt = select(InvoiceModel.id).where(
        InvoiceModel.pelanggan_id == langganan.pelanggan_id,
        InvoiceModel.tgl_jatuh_tempo == langganan.tgl_jatuh_tempo,
    )
    if (await db.execute(existing_invoice_stmt)).scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Invoice untuk periode ini sudah ada.")

    # Handle SQLAlchemy Date untuk formatting
    jatuh_tempo_date = langganan.tgl_jatuh_tempo
    jatuh_tempo_str = safe_format_date(jatuh_tempo_date, "%d/%m/%Y")
    jatuh_tempo_yyyymm = safe_format_date(jatuh_tempo_date, "%Y%m") or "202501"  # Default fallback

    # --- MODIFICATION FOR INVOICE NUMBER ---
    # 1. Sanitize customer name and address
    import re
    nama_pelanggan_singkat = re.sub(r'[^a-zA-Z0-9]', '', pelanggan.nama).upper()
    alamat_singkat = re.sub(r'[^a-zA-Z0-9]', '', pelanggan.alamat or '').upper()[:10]  # Take only first 10 chars
    brand_singkat = re.sub(r'[^a-zA-Z0-9]', '', brand.brand or '').upper()
    # Convert SQLAlchemy Date to Python date for datetime.combine
    from datetime import date, datetime
    jatuh_tempo_python_date = date.fromisoformat(str(langganan.tgl_jatuh_tempo))
    bulan_tahun = datetime.combine(jatuh_tempo_python_date, datetime.min.time()).strftime("%B-%Y").upper()

    # 2. Generate new invoice number in format: BRAND/LAYANAN/NAMA_PELANGGAN/BULAN_TAHUN/ALAMAT_SINGKAT/IDPELANGGAN_LAST3
    nomor_invoice = f"{brand_singkat}/ftth/{nama_pelanggan_singkat}/{bulan_tahun}/{alamat_singkat}/{str(data_teknis.id_pelanggan)[-3:]}"

    # 3. Check for duplicate invoice number and add timestamp if needed
    existing_invoice_number = (await db.execute(
        select(InvoiceModel.id).where(InvoiceModel.invoice_number == nomor_invoice)
    )).scalar_one_or_none()

    if existing_invoice_number:
        # Generate nomor unik dengan tambahan timestamp atau random
        import time
        timestamp = str(int(time.time()))[-6:]  # 6 digit terakhir timestamp
        nomor_invoice = f"{nomor_invoice}/{timestamp}"
    # --- END OF MODIFICATION ---

    # Ambil total harga langsung dari data langganan yang sudah dihitung (prorate + PPN).
    total_harga = float(langganan.harga_awal or 0)
    pajak_persen = float(brand.pajak or 0)

    # Karena Xendit butuh nilai pajak terpisah, kita hitung mundur dari total harga.
    # 1. Cari harga dasar sebelum pajak.
    harga_dasar = total_harga / (1 + (pajak_persen / 100))

    # 2. Hitung nilai pajak berdasarkan selisih total harga dan harga dasar.
    # Gunakan pembulatan untuk konsistensi.
    pajak = round(total_harga - harga_dasar)

    # Pastikan harga dasar untuk item di Xendit juga konsisten.
    harga_dasar = total_harga - pajak

    new_invoice_data = {
        "invoice_number": nomor_invoice,
        "pelanggan_id": pelanggan.id,
        "id_pelanggan": data_teknis.id_pelanggan,
        "brand": brand.brand,
        "total_harga": total_harga,
        "no_telp": pelanggan.no_telp,
        "email": pelanggan.email,
        "tgl_invoice": date.today(),
        "tgl_jatuh_tempo": langganan.tgl_jatuh_tempo,
        "status_invoice": "Belum Dibayar",
    }

    db_invoice = InvoiceModel(**new_invoice_data)
    db.add(db_invoice)
    await db.flush()

    try:
        deskripsi_xendit = ""
        # Handle SQLAlchemy Date untuk formatting
        due_date_obj = db_invoice.tgl_jatuh_tempo
        jatuh_tempo_str_lengkap = safe_format_date(due_date_obj, "%d/%m/%Y")

        if langganan.metode_pembayaran == "Prorate":

            # Hitung harga normal untuk perbandingan
            harga_normal_full = float(paket.harga) * (1 + (float(brand.pajak or 0) / 100))

            # Define invoice_date and due_date before they are referenced
            invoice_date_obj = db_invoice.tgl_invoice
            due_date_obj = db_invoice.tgl_jatuh_tempo

            # Convert ke Python date/datetime dengan aman
            invoice_date = safe_to_datetime(invoice_date_obj) if invoice_date_obj else datetime.now()
            due_date = safe_to_datetime(due_date_obj) if due_date_obj else datetime.now()

            # Cek apakah ini invoice gabungan
            if float(db_invoice.total_harga or 0) > (harga_normal_full + 1):
                # INI TAGIHAN GABUNGAN
                start_day = safe_get_day(invoice_date)
                end_day = safe_get_day(due_date)
                periode_prorate_str = safe_format_date(due_date, "%B %Y")
                due_date_datetime = safe_to_datetime(due_date)
                next_month_date = due_date_datetime + relativedelta(months=1)
                periode_berikutnya_str = safe_format_date(next_month_date.date(), "%B %Y")

                deskripsi_xendit = (
                    f"Biaya internet up to {paket.kecepatan} Mbps. "
                    f"Periode Prorate {start_day}-{end_day} {periode_prorate_str} + "
                    f"Periode {periode_berikutnya_str}"
                )
            else:
                # INI TAGIHAN PRORATE BIASA
                start_day = safe_get_day(invoice_date)
                end_day = safe_get_day(due_date)
                periode_str = safe_format_date(due_date, "%B %Y")
                deskripsi_xendit = (
                    f"Biaya berlangganan internet up to {paket.kecepatan} Mbps, "
                    f"Periode Tgl {start_day}-{end_day} {periode_str}"
                )

        else:  # Otomatis
            deskripsi_xendit = (
                f"Biaya berlangganan internet up to {paket.kecepatan} Mbps "
                f"jatuh tempo pembayaran tanggal {jatuh_tempo_str_lengkap}"
            )

        # Format nomor telepon untuk Xendit (tanpa '+')
        no_telp_bersih = ""
        if pelanggan.no_telp:
            # Hapus spasi dan karakter non-numerik kecuali '+' di awal
            no_telp_bersih = "".join(filter(str.isdigit, pelanggan.no_telp))
            # Handle '0' di depan -> '62'
            if no_telp_bersih.startswith("0"):
                no_telp_bersih = "62" + no_telp_bersih[1:]
            # Jika sudah '62' di depan, biarkan
            elif no_telp_bersih.startswith("62"):
                pass
            # Untuk format lain, coba tambahkan '62' (asumsi nomor lokal tanpa 0)
            else:
                no_telp_bersih = "62" + no_telp_bersih

        no_telp_xendit = no_telp_bersih if no_telp_bersih else None

        # Kirim deskripsi yang sudah dinamis ke Xendit dengan rate limiting
        # Determine priority based on customer type
        priority = InvoicePriority.NORMAL
        if hasattr(pelanggan, 'is_vip') and getattr(pelanggan, 'is_vip', False):
            priority = InvoicePriority.HIGH
        elif hasattr(pelanggan, 'tipe') and getattr(pelanggan, 'tipe', '') == 'bulk':
            priority = InvoicePriority.LOW

        xendit_response = await create_invoice_with_rate_limit(
            invoice=db_invoice,
            pelanggan=pelanggan,
            paket=paket,
            deskripsi_xendit=deskripsi_xendit,
            pajak=pajak,
            no_telp_xendit=no_telp_xendit or "",
            priority=priority
        )

        db_invoice.payment_link = xendit_response.get("short_url", xendit_response.get("invoice_url"))
        db_invoice.xendit_id = xendit_response.get("id")
        db_invoice.xendit_external_id = xendit_response.get("external_id")
        await db.commit()
        await db.refresh(db_invoice)
    except Exception as e:
        await db.rollback()
        logger.error(f"Gagal membuat invoice di Xendit: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Gagal membuat invoice di Xendit: {str(e)}")

    return db_invoice


# POST /invoices/{invoice_id}/mark-as-paid - Tandai invoice lunas manual
# Buat nandai invoice udah dibayar secara manual (bukan via Xendit)
# Path parameters:
# - invoice_id: ID invoice yang mau ditandai lunas
# Request body: metode_pembayaran (cash, transfer, dll)
# Response: data invoice yang udah diupdate
# Permission: butuh permission "edit_invoices"
# Fitur:
# - Update status invoice jadi "Lunas"
# - Update tanggal pembayaran
# - Update status langganan jadi "Aktif"
# - Hitung tanggal jatuh tempo berikutnya
# - Trigger re-aktivasi Mikrotik kalo sebelumnya suspended
# - Kirim notifikasi ke Admin/NOC/Finance
# Validation:
# - Cek invoice harus ada
# - Cek invoice belum lunas sebelumnya
# Error handling: 404 kalo invoice nggak ada, 400 kalo udah lunas
@router.post(
    "/{invoice_id}/mark-as-paid",
    response_model=InvoiceSchema,
    dependencies=[Depends(has_permission("edit_invoices"))],
)
async def mark_invoice_as_paid(invoice_id: int, payload: MarkAsPaidRequest, db: AsyncSession = Depends(get_db)):
    """Menandai sebuah invoice sebagai lunas secara manual."""
    # PERBAIKAN: Eager load semua relasi yang dibutuhkan oleh _process_successful_payment
    # untuk menghindari N+1 query dan membuat proses lebih efisien.
    stmt = (
        select(InvoiceModel)
        .where(InvoiceModel.id == invoice_id)
        .options(
            joinedload(InvoiceModel.pelanggan).options(
                joinedload(PelangganModel.harga_layanan),
                joinedload(PelangganModel.langganan).joinedload(LanggananModel.paket_layanan),
                joinedload(PelangganModel.data_teknis),
            )
        )
    )
    invoice = (await db.execute(stmt)).unique().scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice tidak ditemukan")

    if invoice.status_invoice == "Lunas":
        raise HTTPException(status_code=400, detail="Invoice ini sudah lunas.")

    invoice.metode_pembayaran = payload.metode_pembayaran

    await _process_successful_payment(db, invoice)

    await db.commit()
    await db.refresh(invoice)

    logger.info(f"Invoice {invoice.invoice_number} ditandai lunas secara manual via {payload.metode_pembayaran}")


@router.get("/missing-payment-links", response_model=List[InvoiceSchema])
async def get_invoices_missing_payment_links(
    db: AsyncSession = Depends(get_db),
    current_user: UserModel = Depends(get_current_active_user),
    _: None = Depends(has_permission("view_invoices")),
):
    """
    Get semua invoice yang tidak memiliki payment link dari Xendit.
    Ini berguna untuk finance team agar bisa retry pembuatan payment link.
    """
    stmt = select(InvoiceModel).where(
        InvoiceModel.payment_link.is_(None),
        InvoiceModel.status_invoice == "Belum Dibayar"
    ).order_by(InvoiceModel.tgl_invoice.desc())

    result = await db.execute(stmt)
    invoices = result.scalars().unique().all()

    logger.info(f"Found {len(invoices)} invoices missing payment links")
    return invoices


@router.post("/{invoice_id}/retry-xendit", response_model=dict)
async def retry_invoice_xendit(
    invoice_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: UserModel = Depends(get_current_active_user),
    _: None = Depends(has_permission("edit_invoices")),
):
    """
    Manual retry pembuatan payment link Xendit untuk invoice yang gagal.
    Endpoint ini untuk admin agar bisa retry invoice yang gagal secara manual.

    Parameters:
    - invoice_id: ID invoice yang mau di-retry

    Returns:
    - success: status retry
    - message: pesan hasilnya
    - payment_link: link pembayaran jika berhasil

    Process:
    1. Cek invoice harus ada dan belum punya payment link
    2. Reset retry count ke 0
    3. Coba buat payment link lagi ke Xendit
    4. Update invoice jika berhasil
    """
    from math import floor

    # Load invoice dengan semua relasi yang dibutuhkan
    stmt = (
        select(InvoiceModel)
        .where(InvoiceModel.id == invoice_id)
        .options(
            joinedload(InvoiceModel.pelanggan).options(
                joinedload(PelangganModel.harga_layanan),
                joinedload(PelangganModel.langganan).joinedload(LanggananModel.paket_layanan),
                joinedload(PelangganModel.data_teknis),
            )
        )
    )
    invoice = (await db.execute(stmt)).unique().scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice tidak ditemukan")

    if invoice.xendit_id and invoice.payment_link:
        raise HTTPException(status_code=400, detail="Invoice ini sudah punya payment link")

    if invoice.status_invoice == "Lunas":
        raise HTTPException(status_code=400, detail="Invoice sudah lunas")

    try:
        pelanggan = invoice.pelanggan
        if not pelanggan or not pelanggan.langganan:
            raise HTTPException(status_code=400, detail="Data pelanggan/langganan tidak lengkap")

        paket = pelanggan.langganan[0].paket_layanan
        brand = pelanggan.harga_layanan
        data_teknis = pelanggan.data_teknis

        if not all([paket, brand, data_teknis]):
            raise HTTPException(status_code=400, detail="Data paket/brand/data teknis tidak lengkap")

        # Reset retry count untuk manual retry
        invoice.xendit_retry_count = 0
        invoice.xendit_status = "processing"
        invoice.xendit_last_retry = datetime.now()
        await db.flush()

        # Generate deskripsi yang sama seperti generate_single_invoice
        # Convert SQLAlchemy Date ke Python 
        try:
            jatuh_tempo_date = date.fromisoformat(str(invoice.tgl_jatuh_tempo))
            invoice_date = date.fromisoformat(str(invoice.tgl_invoice))
        except (ValueError, TypeError):
            jatuh_tempo_date = date.today()
            invoice_date = date.today()

        jatuh_tempo_str = datetime.combine(jatuh_tempo_date, datetime.min.time()).strftime("%d/%m/%Y")

        if pelanggan.langganan[0].metode_pembayaran == "Prorate":
            deskripsi_xendit = (
                f"Biaya berlangganan internet up to {paket.kecepatan} Mbps, "
                f"Periode Tgl {invoice_date.day}-{jatuh_tempo_date.day} "
                f"{jatuh_tempo_date.strftime('%B %Y')}"
            )
        else:  # Otomatis
            deskripsi_xendit = (
                f"Biaya berlangganan internet up to {paket.kecepatan} Mbps "
                f"jatuh tempo pembayaran tanggal {jatuh_tempo_str}"
            )

        # Hitung pajak
        pajak_persen = float(brand.pajak)
        harga_dasar = float(paket.harga)
        pajak = floor(harga_dasar * (pajak_persen / 100) + 0.5)

        no_telp_xendit = f"+62{pelanggan.no_telp.lstrip('0')}" if pelanggan.no_telp else ""

        # Coba buat payment link lagi
        xendit_response = await create_invoice_with_rate_limit(
            invoice=invoice,
            pelanggan=pelanggan,
            paket=paket,
            deskripsi_xendit=deskripsi_xendit,
            pajak=pajak,
            no_telp_xendit=no_telp_xendit,
            priority=InvoicePriority.HIGH  # High priority untuk manual retry
        )

        # Validasi response
        if not xendit_response or not xendit_response.get("id"):
            raise ValueError(f"Invalid Xendit response: {xendit_response}")

        # Update invoice dengan payment link
        invoice.payment_link = xendit_response.get("short_url", xendit_response.get("invoice_url"))
        invoice.xendit_id = xendit_response.get("id")
        invoice.xendit_external_id = xendit_response.get("external_id")
        invoice.xendit_status = "completed"
        invoice.xendit_error_message = None

        await db.commit()

        logger.info(f"✅ Manual retry SUCCESS: Invoice {invoice.invoice_number} - {pelanggan.nama}")
        logger.info(f"📱 Payment link: {invoice.payment_link}")

        return {
            "success": True,
            "message": f"Payment link berhasil dibuat untuk invoice {invoice.invoice_number}",
            "payment_link": invoice.payment_link,
            "xendit_id": invoice.xendit_id,
            "pelanggan": pelanggan.nama
        }

    except Exception as e:
        await db.rollback()

        # Update error tracking
        invoice.xendit_retry_count += 1
        invoice.xendit_status = "failed"
        invoice.xendit_error_message = str(e)
        await db.commit()

        logger.error(f"❌ Manual retry FAILED: Invoice {invoice.invoice_number} - {str(e)}")

        raise HTTPException(
            status_code=500,
            detail=f"Gagal membuat payment link: {str(e)}"
        )


@router.post("/batch-retry-xendit", response_model=dict)
async def batch_retry_failed_invoices(
    db: AsyncSession = Depends(get_db),
    current_user: UserModel = Depends(get_current_active_user),
    _: None = Depends(has_permission("edit_invoices")),
):
    """
    Batch retry untuk semua invoice yang gagal (belum ada payment link).
    Endpoint ini untuk admin agar bisa retry semua invoice gagal sekaligus.

    Returns:
    - total_processed: total invoice yang diproses
    - success_count: jumlah yang berhasil
    - failed_count: jumlah yang gagal
    - results: detail hasil per invoice
    """
    from math import floor

    # Cari semua invoice yang belum ada payment link-nya
    stmt = (
        select(InvoiceModel)
        .where(
            InvoiceModel.payment_link.is_(None),
            InvoiceModel.status_invoice == "Belum Dibayar"
        )
        .options(
            joinedload(InvoiceModel.pelanggan).options(
                joinedload(PelangganModel.harga_layanan),
                joinedload(PelangganModel.langganan).joinedload(LanggananModel.paket_layanan),
                joinedload(PelangganModel.data_teknis),
            )
        )
        .order_by(InvoiceModel.created_at.desc())
        .limit(50)  # Batasi max 50 invoice per batch
    )

    invoices = (await db.execute(stmt)).unique().scalars().all()

    if not invoices:
        return {
            "success": True,
            "message": "Tidak ada invoice yang perlu di-retry",
            "total_processed": 0,
            "success_count": 0,
            "failed_count": 0,
            "results": []
        }

    total_processed = len(invoices)
    success_count = 0
    failed_count = 0
    results = []

    logger.info(f"🔄 Starting batch retry for {total_processed} invoices by user {current_user.name}")

    for invoice in invoices:
        try:
            pelanggan = invoice.pelanggan
            if not pelanggan or not pelanggan.langganan:
                results.append({
                    "invoice_id": invoice.id,
                    "invoice_number": invoice.invoice_number,
                    "success": False,
                    "message": "Data pelanggan/langganan tidak lengkap"
                })
                failed_count += 1
                continue

            paket = pelanggan.langganan[0].paket_layanan
            brand = pelanggan.harga_layanan

            if not all([paket, brand]):
                results.append({
                    "invoice_id": invoice.id,
                    "invoice_number": invoice.invoice_number,
                    "success": False,
                    "message": "Data paket/brand tidak lengkap"
                })
                failed_count += 1
                continue

            # Reset retry count
            invoice.xendit_retry_count = 0
            invoice.xendit_status = "processing"
            invoice.xendit_last_retry = datetime.now()
            await db.flush()

            # Generate deskripsi
            # Convert SQLAlchemy Date ke Python date dengan aman
            try:
                jatuh_tempo_date = date.fromisoformat(str(invoice.tgl_jatuh_tempo))
                invoice_date = date.fromisoformat(str(invoice.tgl_invoice))
            except (ValueError, TypeError):
                jatuh_tempo_date = date.today()
                invoice_date = date.today()

            jatuh_tempo_str = datetime.combine(jatuh_tempo_date, datetime.min.time()).strftime("%d/%m/%Y")

            if pelanggan.langganan[0].metode_pembayaran == "Prorate":
                deskripsi_xendit = (
                    f"Biaya berlangganan internet up to {paket.kecepatan} Mbps, "
                    f"Periode Tgl {invoice_date.day}-{jatuh_tempo_date.day} "
                    f"{jatuh_tempo_date.strftime('%B %Y')}"
                )
            else:
                deskripsi_xendit = (
                    f"Biaya berlangganan internet up to {paket.kecepatan} Mbps "
                    f"jatuh tempo pembayaran tanggal {jatuh_tempo_str}"
                )

            # Hitung pajak
            pajak_persen = float(brand.pajak)
            harga_dasar = float(paket.harga)
            pajak = floor(harga_dasar * (pajak_persen / 100) + 0.5)

            no_telp_xendit = f"+62{pelanggan.no_telp.lstrip('0')}" if pelanggan.no_telp else ""

            # Coba buat payment link
            xendit_response = await create_invoice_with_rate_limit(
                invoice=invoice,
                pelanggan=pelanggan,
                paket=paket,
                deskripsi_xendit=deskripsi_xendit,
                pajak=pajak,
                no_telp_xendit=no_telp_xendit,
                priority=InvoicePriority.NORMAL
            )

            if xendit_response and xendit_response.get("id"):
                # Update invoice
                invoice.payment_link = xendit_response.get("short_url", xendit_response.get("invoice_url"))
                invoice.xendit_id = xendit_response.get("id")
                invoice.xendit_external_id = xendit_response.get("external_id")
                invoice.xendit_status = "completed"
                invoice.xendit_error_message = None

                results.append({
                    "invoice_id": invoice.id,
                    "invoice_number": invoice.invoice_number,
                    "success": True,
                    "message": "Payment link berhasil dibuat",
                    "payment_link": invoice.payment_link,
                    "pelanggan": pelanggan.nama
                })
                success_count += 1

                logger.info(f"✅ Batch retry SUCCESS: {invoice.invoice_number} - {pelanggan.nama}")
            else:
                raise ValueError("Invalid Xendit response")

        except Exception as e:
            # Update error tracking
            invoice.xendit_retry_count += 1
            invoice.xendit_status = "failed"
            invoice.xendit_error_message = str(e)

            results.append({
                "invoice_id": invoice.id,
                "invoice_number": invoice.invoice_number,
                "success": False,
                "message": str(e)
            })
            failed_count += 1

            logger.error(f"❌ Batch retry FAILED: {invoice.invoice_number} - {str(e)}")

    await db.commit()

    logger.info(f"🏁 Batch retry completed: {success_count} success, {failed_count} failed")

    return {
        "success": True,
        "message": f"Batch retry selesai: {success_count} berhasil, {failed_count} gagal",
        "total_processed": total_processed,
        "success_count": success_count,
        "failed_count": failed_count,
        "results": results
    }


# DELETE /invoices/{invoice_id} - Hapus invoice
# Buat hapus invoice dari sistem
# Path parameters:
# - invoice_id: ID invoice yang mau dihapus
# Response: 204 No Content (sukses tapi nggak ada response body)
# Permission: butuh permission "delete_invoices"
# Warning: HATI-HATI! Ini akan hapus invoice permanen
# Note: Payment link di Xendit mungkin masih aktif
# Error handling: 404 kalau invoice nggak ketemu
@router.delete(
    "/{invoice_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(has_permission("delete_invoices"))],
)
async def delete_invoice(invoice_id: int, db: AsyncSession = Depends(get_db)):
    """Menghapus satu invoice berdasarkan ID-nya."""

    db_invoice = await db.get(InvoiceModel, invoice_id)

    if not db_invoice:
        raise HTTPException(status_code=404, detail="Invoice tidak ditemukan")

    await db.delete(db_invoice)
    await db.commit()

    return None


# Logic atau EndPoint untuk melihat Status Invoice itu masih layar atau sudah kadaluarsa ??


# POST /invoices/internal/update-overdue-status - Update status invoice kadaluarsa (Internal)
# Endpoint internal buat update status invoice yang telat bayar dan suspend layanan
# Endpoint ini HARUS dipanggil oleh scheduler (cron job) setiap hari
# Response: jumlah invoice yang diupdate + jumlah layanan yang disuspend
# Security: include_in_schema=False (nggak muncul di dokumentasi API publik)
# Logic:
# - Cari invoice "Belum Dibayar" yang udah lewat 5 hari dari jatuh tempo
# - Update status jadi "Kadaluarsa"
# - Update status langganan jadi "Suspended"
# - Trigger suspend Mikrotik
# Aturan: Kadaluarsa kalo hari ini adalah hari ke-6 setelah jatuh tempo (lewat 5 hari)
# Use case: automated system untuk penagihan dan penonaktifan layanan
# Error handling: graceful error handling, log semua error
@router.post(
    "/internal/update-overdue-status",
    status_code=status.HTTP_200_OK,
    include_in_schema=False,
)
async def update_overdue_invoices(db: AsyncSession = Depends(get_db)):
    """
    Endpoint internal untuk memperbarui status invoice menjadi 'Kadaluarsa' dan men-suspend layanan.
    Endpoint ini HARUS dipanggil oleh scheduler (cron job) setiap hari dan tidak boleh terekspos ke publik.
    `include_in_schema=False` digunakan agar tidak muncul di dokumentasi API publik.
    """
    today = date.today()
    # Aturan: Kadaluarsa jika hari ini adalah hari ke-6 setelah jatuh tempo (lewat 5 hari).
    overdue_threshold_date = today - timedelta(days=5)

    logger = logging.getLogger("app.routers.invoice")
    logger.info("Scheduler-triggered job: Mencari invoice kadaluarsa...")

    # 1. Cari semua invoice yang 'Belum Dibayar' dan sudah melewati masa tenggang
    stmt = (
        select(InvoiceModel)
        # PERBAIKAN: Eager load data_teknis untuk mencegah N+1 query di dalam loop.
        .options(
            joinedload(InvoiceModel.pelanggan).options(
                joinedload(PelangganModel.langganan),
                joinedload(PelangganModel.data_teknis),
            )
        ).where(
            and_(
                InvoiceModel.status_invoice == "Belum Dibayar",
                InvoiceModel.tgl_jatuh_tempo < overdue_threshold_date,
            )
        )
    )

    # FIX: Tambahkan .unique() untuk collection eager loading
    overdue_invoices = (await db.execute(stmt)).scalars().unique().all()

    if not overdue_invoices:
        logger.info("Tidak ada invoice kadaluarsa yang perlu diperbarui.")
        return {"message": "Tidak ada invoice kadaluarsa yang perlu diperbarui."}

    updated_count = 0
    suspended_count = 0
    # 2. Update status dan suspend layanan untuk setiap invoice yang ditemukan
    for invoice in overdue_invoices:
        invoice.status_invoice = "Kadaluarsa"
        db.add(invoice)
        updated_count += 1

        # Logika untuk men-suspend layanan pelanggan
        try:
            pelanggan = invoice.pelanggan
            if pelanggan and pelanggan.langganan:
                langganan_pelanggan = pelanggan.langganan[0]
                if langganan_pelanggan.status != "Suspended":
                    langganan_pelanggan.status = "Suspended"
                    db.add(langganan_pelanggan)

                    # PERBAIKAN: Panggil trigger_mikrotik_update dengan argumen yang benar.
                    data_teknis = pelanggan.data_teknis
                    if data_teknis:
                        await mikrotik_service.trigger_mikrotik_update(
                            db,
                            langganan_pelanggan,
                            data_teknis,
                            data_teknis.id_pelanggan,
                        )
                        suspended_count += 1
                        logger.info(f"Layanan untuk {pelanggan.nama} (Invoice: {invoice.invoice_number}) telah di-suspend.")
                    else:
                        logger.error(
                            f"Data Teknis tidak ditemukan untuk pelanggan {pelanggan.nama}, suspend di Mikrotik dilewati."
                        )
        except Exception as e:
            logger.error(f"Gagal men-suspend layanan untuk invoice {invoice.invoice_number}: {e}")

    await db.commit()

    message = f"Proses selesai. {updated_count} invoice diperbarui menjadi 'Kadaluarsa'. {suspended_count} layanan di-suspend."
    logger.info(message)
    return {"message": message}


# GET /invoices/export-payment-links-excel - Export payment links ke Excel
# Buat export semua payment link invoice ke file Excel
# Query parameters:
# - search: filter pencarian (sama seperti di list)
# - status_invoice: filter berdasarkan status
# - start_date: filter tanggal mulai
# - end_date: filter tanggal akhir
# Response: file Excel (.xlsx) dengan kolom:
#   ID Invoice, Nomor Invoice, Nama Pelanggan, ID Pelanggan, Alamat,
#   Total Harga, Status Invoice, Tanggal Invoice, Tanggal Jatuh Tempo,
#   Payment Link, Email, No. Telepon, Brand
# Use case: buat share payment links ke tim collection atau customer
# Performance: query dengan eager loading biar efficient
# Format: Excel dengan header bold dan auto-size columns
@router.get("/export-payment-links-excel")
async def export_payment_links_excel(
    db: AsyncSession = Depends(get_db),
    search: Optional[str] = None,
    status_invoice: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
):
    """Export payment links dari invoice ke file Excel."""
    query = (
        select(InvoiceModel)
        .join(InvoiceModel.pelanggan)
        .options(joinedload(InvoiceModel.pelanggan).joinedload(PelangganModel.harga_layanan))
    )

    if search:
        search_term = f"%{search}%"
        query = query.where(
            or_(
                InvoiceModel.invoice_number.ilike(search_term),
                PelangganModel.nama.ilike(search_term),
                InvoiceModel.id_pelanggan.ilike(search_term),
            )
        )

    if status_invoice:
        query = query.where(InvoiceModel.status_invoice == status_invoice)

    if start_date:
        query = query.where(InvoiceModel.tgl_jatuh_tempo >= start_date)
    if end_date:
        query = query.where(InvoiceModel.tgl_jatuh_tempo <= end_date)

    # Hanya ambil invoice dengan payment_link
    query = query.where(InvoiceModel.payment_link.isnot(None))

    result = await db.execute(query)
    # FIX: Tambahkan .unique() untuk collection eager loading
    invoices = result.scalars().unique().all()

    # Buat workbook dan worksheet pertama untuk Payment Links
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Payment Links Invoice"

    # Definisikan header
    headers = [
        "ID Invoice",
        "Nomor Invoice",
        "Nama Pelanggan",
        "ID Pelanggan",
        "Alamat Pelanggan",
        "Total Harga",
        "Status Invoice",
        "Tanggal Invoice",
        "Tanggal Jatuh Tempo",
        "Payment Link",
        "Email",
        "No. Telepon",
        "Brand",
    ]

    # Tambahkan header ke worksheet (dengan null check)
    if ws is not None:
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            # Gunakan import langsung untuk styles
            from openpyxl.styles import Font, PatternFill, Alignment

            if cell is not None:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Isi data (dengan null check)
    if ws is not None:
        for row_num, invoice in enumerate(invoices, 2):
            ws.cell(row=row_num, column=1, value=invoice.id)
            ws.cell(row=row_num, column=2, value=invoice.invoice_number)
            ws.cell(row=row_num, column=3, value=invoice.pelanggan.nama if invoice.pelanggan else "")
            ws.cell(row=row_num, column=4, value=invoice.id_pelanggan)
            ws.cell(row=row_num, column=5, value=invoice.pelanggan.alamat if invoice.pelanggan else "")
            ws.cell(row=row_num, column=6, value=float(invoice.total_harga) if invoice.total_harga else 0)
            ws.cell(row=row_num, column=7, value=invoice.status_invoice)

            # Handle SQLAlchemy Date untuk Excel export
            invoice_date = invoice.tgl_invoice
            due_date = invoice.tgl_jatuh_tempo

            ws.cell(row=row_num, column=8, value=safe_format_date(invoice_date, "%Y-%m-%d"))
            ws.cell(row=row_num, column=9, value=safe_format_date(due_date, "%Y-%m-%d"))

            ws.cell(row=row_num, column=10, value=invoice.payment_link)
            ws.cell(row=row_num, column=11, value=invoice.email or "")
            ws.cell(row=row_num, column=12, value=invoice.no_telp or "")
            ws.cell(row=row_num, column=13, value=invoice.brand or "")

        # Auto-adjust column width (dengan null check)
        from openpyxl.utils import get_column_letter

        if ws is not None:
            for column in ws.columns:
                max_length = 0
                # Handle column[0].column yang mungkin None
                first_cell = column[0] if column else None
                if first_cell and hasattr(first_cell, "column") and first_cell.column is not None:
                    column_letter = get_column_letter(first_cell.column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    if ws is not None:
                        ws.column_dimensions[column_letter].width = adjusted_width

    # Buat sheet kedua untuk Matrix Persentase Invoice
    ws_matrix = wb.create_sheet("Matrix Persentase Invoice")
    if ws_matrix is not None:
        # Hitung statistik dari invoices yang sudah difilter
        total_invoices = len(invoices)
        if total_invoices > 0:
            # Hitung berdasarkan status invoice
            lunas_count = len([inv for inv in invoices if inv.status_invoice == 'Lunas'])
            belum_dibayar_count = len([inv for inv in invoices if inv.status_invoice == 'Belum Dibayar'])
            kadaluarsa_count = len([inv for inv in invoices if inv.status_invoice == 'Kadaluarsa'])

            # Hitung persentase
            lunas_percent = (lunas_count / total_invoices) * 100
            belum_dibayar_percent = (belum_dibayar_count / total_invoices) * 100
            kadaluarsa_percent = (kadaluarsa_count / total_invoices) * 100

            # Format tanggal untuk header
            start_date_str = start_date.strftime('%d/%m/%Y') if start_date else 'Awal'
            end_date_str = end_date.strftime('%d/%m/%Y') if end_date else 'Sekarang'

            # Styling untuk header matrix
            header_font = Font(bold=True, size=14)
            title_font = Font(bold=True, size=12)
            data_font = Font(bold=False, size=11)
            blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            green_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
            red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

            # Header Matrix
        ws_matrix.merge_cells('A1:D1')
        ws_matrix.cell(row=1, column=1, value="MATRIX LAPORAN PERSENTASE INVOICE")
        ws_matrix.cell(row=1, column=1).font = header_font
        ws_matrix.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        # Informasi Periode dan Total
        ws_matrix.cell(row=2, column=1, value=f"Periode: {start_date_str} - {end_date_str}")
        ws_matrix.cell(row=2, column=1).font = title_font
        ws_matrix.merge_cells('A2:D2')

        ws_matrix.cell(row=3, column=1, value=f"Total Invoice (Filter Aktif): {total_invoices}")
        ws_matrix.cell(row=3, column=1).font = title_font
        ws_matrix.merge_cells('A3:D3')

        # Spasi
        ws_matrix.cell(row=4, column=1, value="")

        # Header Tabel
        ws_matrix.cell(row=5, column=1, value="Status Invoice")
        ws_matrix.cell(row=5, column=2, value="Jumlah")
        ws_matrix.cell(row=5, column=3, value="Persentase")
        ws_matrix.cell(row=5, column=4, value="Visual")

        # Styling header tabel
        for col in range(1, 5):
            cell = ws_matrix.cell(row=5, column=col)
            cell.font = title_font
            cell.fill = gray_fill
            cell.alignment = Alignment(horizontal='center')

        # Data Total Invoice
        ws_matrix.cell(row=6, column=1, value="Total Invoice")
        ws_matrix.cell(row=6, column=2, value=total_invoices)
        ws_matrix.cell(row=6, column=3, value="100.0%")
        ws_matrix.cell(row=6, column=4, value="████████████████████")
        for col in range(1, 5):
            cell = ws_matrix.cell(row=6, column=col)
            cell.font = data_font
            cell.fill = blue_fill

        # Data Invoice Lunas
        ws_matrix.cell(row=7, column=1, value="Invoice Lunas")
        ws_matrix.cell(row=7, column=2, value=lunas_count)
        ws_matrix.cell(row=7, column=3, value=f"{lunas_percent:.1f}%")
        # Visual bar chart dengan text blocks
        bar_length = int(lunas_percent / 5)  # 1 block = 5%
        ws_matrix.cell(row=7, column=4, value="█" * bar_length)
        for col in range(1, 4):
            cell = ws_matrix.cell(row=7, column=col)
            cell.font = data_font
            cell.fill = green_fill

        # Data Invoice Belum Dibayar
        ws_matrix.cell(row=8, column=1, value="Invoice Belum Dibayar")
        ws_matrix.cell(row=8, column=2, value=belum_dibayar_count)
        ws_matrix.cell(row=8, column=3, value=f"{belum_dibayar_percent:.1f}%")
        bar_length = int(belum_dibayar_percent / 5)
        ws_matrix.cell(row=8, column=4, value="█" * bar_length)
        for col in range(1, 4):
            cell = ws_matrix.cell(row=8, column=col)
            cell.font = data_font
            cell.fill = yellow_fill

        # Data Invoice Kadaluarsa
        ws_matrix.cell(row=9, column=1, value="Invoice Kadaluarsa")
        ws_matrix.cell(row=9, column=2, value=kadaluarsa_count)
        ws_matrix.cell(row=9, column=3, value=f"{kadaluarsa_percent:.1f}%")
        bar_length = int(kadaluarsa_percent / 5)
        ws_matrix.cell(row=9, column=4, value="█" * bar_length)
        for col in range(1, 4):
            cell = ws_matrix.cell(row=9, column=col)
            cell.font = data_font
            cell.fill = red_fill

        # Spasi
        ws_matrix.cell(row=10, column=1, value="")

        # Insight/Kesimpulan
        ws_matrix.cell(row=11, column=1, value="INSIGHT & KESIMPULAN")
        ws_matrix.cell(row=11, column=1).font = title_font
        ws_matrix.merge_cells('A11:D11')

        # Analisis pembayaran
        if lunas_percent >= 70:
            payment_insight = f"Tingkat pembayaran sangat baik ({lunas_percent:.1f}% lunas)"
        elif lunas_percent >= 50:
            payment_insight = f"Tingkat pembayaran cukup baik ({lunas_percent:.1f}% lunas)"
        else:
            payment_insight = f"Tingkat pembayaran perlu ditingkatkan ({lunas_percent:.1f}% lunas)"

        ws_matrix.cell(row=12, column=1, value=payment_insight)
        ws_matrix.merge_cells('A12:D12')

        # Rekomendasi tindakan
        if kadaluarsa_percent > 20:
            rekomendasi = "Perlu follow-up intensif untuk invoice kadaluarsa"
        elif belum_dibayar_percent > 50:
            rekomendasi = "Perlu reminder rutin untuk pembayaran"
        else:
            rekomendasi = "Status pembayaran dalam kondisi normal"

        ws_matrix.cell(row=13, column=1, value=rekomendasi)
        ws_matrix.merge_cells('A13:D13')

        # Auto-adjust column width untuk sheet matrix
        for column in ws_matrix.columns:
            max_length = 0
            first_cell = column[0] if column else None
            if first_cell and hasattr(first_cell, "column") and first_cell.column is not None:
                column_letter = get_column_letter(first_cell.column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_matrix.column_dimensions[column_letter].width = adjusted_width

    else:
        # Jika tidak ada data, gunakan styling dasar
        header_font = Font(bold=True, size=14)
        ws_matrix.cell(row=1, column=1, value="MATRIX LAPORAN PERSENTASE INVOICE")
        ws_matrix.cell(row=1, column=1).font = header_font
        ws_matrix.merge_cells('A1:D1')

        ws_matrix.cell(row=2, column=1, value="Tidak ada data invoice yang memenuhi filter")
        ws_matrix.merge_cells('A2:D2')

    # Simpan workbook ke BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Kembalikan file sebagai response
    return Response(
        buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=payment_links_invoice.xlsx"},
    )


# GET /invoices/count - Hitung total jumlah invoice
# Buat ngambil total jumlah invoice di database dengan optional filter
# Query parameters:
# - status_invoice: filter berdasarkan status (Belum Dibayar, Lunas, Kadaluarsa)
# - start_date: filter tanggal jatuh tempo mulai dari
# - end_date: filter tanggal jatuh tempo sampai dengan
# - search: cari berdasarkan nomor invoice, nama pelanggan, atau ID pelanggan
# Response: integer (total count)
# Use case: buat dashboard atau statistik
# Performance: simple count query dengan filter yang sama seperti list endpoint
@router.get("/count", response_model=int)
async def get_invoice_count(
    db: AsyncSession = Depends(get_db),
    search: Optional[str] = None,
    status_invoice: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
):
    """
    Menghitung total jumlah invoice dengan filter opsional.
    """
    count_query = select(func.count(InvoiceModel.id))

    # Join dengan pelanggan untuk pencarian
    count_query = count_query.join(InvoiceModel.pelanggan)

    if search:
        search_term = f"%{search}%"
        count_query = count_query.where(
            or_(
                InvoiceModel.invoice_number.ilike(search_term),
                PelangganModel.nama.ilike(search_term),
                InvoiceModel.id_pelanggan.ilike(search_term),
            )
        )

    if status_invoice:
        count_query = count_query.where(InvoiceModel.status_invoice == status_invoice)

    if start_date:
        count_query = count_query.where(InvoiceModel.tgl_jatuh_tempo >= start_date)
    if end_date:
        count_query = count_query.where(InvoiceModel.tgl_jatuh_tempo <= end_date)

    result = await db.execute(count_query)
    total_count = result.scalar_one()
    return total_count


# ====================================================================
# MONITORING & FIX ENDPOINTS - PRODUCTION RELIABILITY
# ====================================================================

@router.get("/generation-status")
async def get_invoice_generation_status(
    db: AsyncSession = Depends(get_db),
    current_user: UserModel = Depends(get_current_active_user)
):
    """Simple dashboard untuk monitor invoice generation"""

    today = date.today()

    # Total invoice hari ini
    total_today = await db.execute(
        select(func.count(InvoiceModel.id))
        .where(InvoiceModel.tgl_invoice == today)
    )

    # Invoice dengan payment link
    with_payment_link = await db.execute(
        select(func.count(InvoiceModel.id))
        .where(
            InvoiceModel.tgl_invoice == today,
            InvoiceModel.payment_link.isnot(None)
        )
    )

    total = total_today.scalar() or 0
    success = with_payment_link.scalar() or 0
    failed = total - success

    return {
        "date": today.isoformat(),
        "total_invoices": total,
        "successful_payments": success,
        "failed_payments": failed,
        "success_rate": round((success / total * 100) if total > 0 else 0, 1),
        "status": "HEALTHY" if failed == 0 else "NEEDS_ATTENTION" if failed <= 3 else "CRITICAL"
    }


@router.post("/fix-missing-payment-links")
async def fix_missing_payment_links(
    limit: int = 20,
    db: AsyncSession = Depends(get_db),
    current_user: UserModel = Depends(get_current_active_user)
):
    """Fix invoices yang tidak punya payment link"""

    from ..services.rate_limiter import create_invoice_with_rate_limit

    # Cari invoice tanpa payment link
    invoices_to_fix = await db.execute(
        select(InvoiceModel)
        .where(InvoiceModel.payment_link.is_(None))
        .order_by(InvoiceModel.created_at.desc())
        .limit(limit)
        .options(
            selectinload(InvoiceModel.pelanggan).selectinload(PelangganModel.harga_layanan),
            selectinload(InvoiceModel.pelanggan).selectinload(PelangganModel.langganan).selectinload(LanggananModel.paket_layanan),
        )
    )

    fixed_count = 0
    failed_count = 0

    for invoice in invoices_to_fix.scalars().all():
        try:
            pelanggan = invoice.pelanggan
            if not pelanggan or not pelanggan.langganan:
                continue

            paket = pelanggan.langganan[0].paket_layanan

            # Generate payment link
            xendit_response = await create_invoice_with_rate_limit(
                invoice=invoice,
                pelanggan=pelanggan,
                paket=paket,
                deskripsi_xendit=f"Invoice Payment - {invoice.invoice_number}",
                pajak=float(invoice.total_harga) - float(paket.harga) if paket else 0,
                no_telp_xendit=f"+62{pelanggan.no_telp.lstrip('0')}" if pelanggan.no_telp else ""
            )

            # Update invoice
            invoice.payment_link = xendit_response.get("short_url", xendit_response.get("invoice_url"))
            invoice.xendit_id = xendit_response.get("id")
            invoice.xendit_external_id = xendit_response.get("external_id")

            fixed_count += 1

        except Exception as e:
            failed_count += 1
            logger.error(f"Failed to fix invoice {invoice.invoice_number}: {e}")

    await db.commit()

    return {
        "message": f"Fixed {fixed_count} invoices, {failed_count} failed",
        "fixed_count": fixed_count,
        "failed_count": failed_count
    }
